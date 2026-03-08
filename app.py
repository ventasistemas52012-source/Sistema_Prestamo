from datetime import datetime
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from openpyxl.styles import Font, Alignment, Border, Side
from flask import Flask, render_template, request, redirect, url_for
from openpyxl.drawing.image import Image
import sqlite3
from openpyxl.styles import Border, Side, Font, Alignment
from datetime import datetime, timedelta
MORA_DIARIA = 2  # puedes cambiar el valor si quieres
from openpyxl import Workbook
from flask import send_file
import os
from datetime import date
import os
from werkzeug.utils import secure_filename
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "prestamos.db")



app = Flask(__name__)

UPLOAD_FOLDER = "static/uploads"
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)


from datetime import date



app.secret_key = 'clave_super_secreta_123'

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"

class User(UserMixin):
    def __init__(self, id):
        self.id = id

# Usuario fijo (luego podemos hacerlo con base de datos)
usuarios = {
    "admin": {"password": "1234"}
}

@login_manager.user_loader
def load_user(user_id):
    return User(user_id)


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        if username in usuarios and usuarios[username]['password'] == password:
            user = User(username)
            login_user(user)
            return redirect(url_for('index'))
        else:
            return "Usuario o contraseña incorrectos"

    return render_template('login.html')



@app.route("/dashboard")
@login_required
def dashboard():
    return render_template("inicio.html")

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))



# Crear base de datos si no existe
def init_db():
    conn = sqlite3.connect('prestamos.db')
    cursor = conn.cursor()

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS clientes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nombre TEXT,
        apellidos TEXT,
        dni TEXT,
        direccion TEXT,
        telefono TEXT,
        monto REAL,
        interes REAL,
        total REAL,
        cuotas INTEGER,
        tipo_pago TEXT,
        dni_frontal TEXT,
        dni_reverso TEXT,
        foto_rostro TEXT,
        recibo_servicio TEXT
    )
    """)
    
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS cronograma (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cliente_id INTEGER,
        fecha_pago TEXT,
        cuota REAL,
        estado TEXT
    )
    """)
    
    conn.commit()
    conn.close()

init_db()

@app.route("/")
def index():

    conn = sqlite3.connect("DB_PATH")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    # CLIENTES
    cursor.execute("SELECT * FROM clientes")
    clientes = cursor.fetchall()

    # CRONOGRAMA
    cursor.execute("SELECT fecha_pago, estado FROM cronograma")
    cuotas = cursor.fetchall()

    hoy = datetime.today().date()

    contador_vencidos = 0

    for cuota in cuotas:

        fecha_pago = datetime.strptime(cuota["fecha_pago"], "%d/%m/%Y").date()

        if hoy > fecha_pago and not cuota["estado"].startswith("Pagado"):
            contador_vencidos += 1

    conn.close()

    return render_template(
        "index.html",
        clientes=clientes,
        contador_vencidos=contador_vencidos
    )


@app.route("/nuevo", methods=["GET", "POST"])
def nuevo():
    if request.method == "POST":
        nombre = request.form["nombre"]
        apellidos = request.form["apellidos"]
        dni = request.form["dni"]
        direccion = request.form["direccion"]
        telefono = request.form["telefono"]
        monto = float(request.form["monto"])
        interes = float(request.form["interes"])
        tipo_pago = request.form["tipo_pago"]
        cuotas = int(request.form["cuotas"])
        dni_frontal = request.files.get('dni_frontal')
        dni_reverso = request.files.get('dni_reverso')
        foto_rostro = request.files.get('foto_rostro')
        recibo_servicio = request.files.get('recibo_servicio')

        # Guardar archivos
        dni_frontal_filename = secure_filename(dni_frontal.filename)
        dni_reverso_filename = secure_filename(dni_reverso.filename)
        foto_rostro_filename = secure_filename(foto_rostro.filename)
        recibo_servicio_filename = secure_filename(recibo_servicio.filename)

        dni_frontal.save(os.path.join(app.config["UPLOAD_FOLDER"], dni_frontal_filename))
        dni_reverso.save(os.path.join(app.config["UPLOAD_FOLDER"], dni_reverso_filename))
        foto_rostro.save(os.path.join(app.config["UPLOAD_FOLDER"], foto_rostro_filename))
        recibo_servicio.save(os.path.join(app.config["UPLOAD_FOLDER"], recibo_servicio_filename))
        
        total = monto + (monto * interes / 100)
        valor_cuota = round(total / cuotas, 2)
        
        conn = sqlite3.connect("DB.PATH")
        cursor = conn.cursor()

        cursor.execute("""
        INSERT INTO clientes
        (nombre, apellidos, dni, direccion, telefono, monto, interes, total, tipo_pago, cuotas,
         dni_frontal, dni_reverso, foto_rostro, recibo_servicio)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            nombre,
            apellidos,
            dni,
            direccion,
            telefono,
            monto,
            interes,
            total,
            tipo_pago,
            cuotas,
            dni_frontal_filename,
            dni_reverso_filename,
            foto_rostro_filename,
            recibo_servicio_filename
        ))

        cliente_id = cursor.lastrowid
        
        fecha_actual = datetime.today()
        
        for i in range(cuotas):
            if tipo_pago == "semanal":
                fecha_pago = fecha_actual + timedelta(weeks=i+1)
            elif tipo_pago == "quincenal":
                fecha_pago = fecha_actual + timedelta(days=15*(i+1))
            else:
                fecha_pago = fecha_actual + timedelta(days=30*(i+1))
            
            cursor.execute("""
            INSERT INTO cronograma (cliente_id, fecha_pago, cuota, estado)
            VALUES (?, ?, ?, ?)
            """, (cliente_id, fecha_pago.strftime("%d/%m/%Y"), valor_cuota, "Pendiente"))
        
        conn.commit()
        conn.close()
        
        return redirect("/")
    
    return render_template("nuevo.html")


@app.route("/editar/<int:cliente_id>", methods=["GET", "POST"])
@login_required
def editar(cliente_id):

    conn = sqlite3.connect("DB.PATH")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    if request.method == "POST":

        nombre = request.form["nombre"]
        apellidos = request.form["apellidos"]
        dni = request.form["dni"]
        direccion = request.form["direccion"]
        telefono = request.form["telefono"]
        monto = float(request.form["monto"])
        interes = float(request.form["interes"])
        cuotas = int(request.form["cuotas"])
        tipo_pago = request.form["tipo_pago"]

        foto = request.files.get('foto_rostro')
        dni_frontal = request.files.get('dni_frontal')
        dni_reverso = request.files.get('dni_reverso')
        recibo_servicio = request.files.get('recibo_servicio')
        eliminar_foto = request.form.get('eliminar_foto')

        if eliminar_foto:
            cursor.execute("UPDATE clientes SET foto_rostro=NULL WHERE id=?", (cliente_id,))


        if foto and foto.filename != "":
            nombre_archivo = secure_filename(foto.filename)
            ruta = os.path.join("static/uploads", nombre_archivo)
            foto.save(ruta)

            cursor.execute("UPDATE clientes SET foto_rostro=? WHERE id=?", (nombre_archivo, cliente_id))
            conn.commit()

        if dni_frontal and dni_frontal.filename != "":
            nombre_archivo = secure_filename(dni_frontal.filename)
            ruta = os.path.join("static/uploads", nombre_archivo)
            dni_frontal.save(ruta)

            cursor.execute(
                "UPDATE clientes SET dni_frontal=? WHERE id=?",
                (nombre_archivo, cliente_id)
            )

        if dni_reverso and dni_reverso.filename != "":
            nombre_archivo = secure_filename(dni_reverso.filename)
            ruta = os.path.join("static/uploads", nombre_archivo)
            dni_reverso.save(ruta)

            cursor.execute(
                "UPDATE clientes SET dni_reverso=? WHERE id=?",
                (nombre_archivo, cliente_id)
            )

        if recibo_servicio and recibo_servicio.filename != "":
            nombre_archivo = secure_filename(recibo_servicio.filename)
            ruta = os.path.join("static/uploads", nombre_archivo)
            recibo_servicio.save(ruta)

            cursor.execute(
                "UPDATE clientes SET recibo_servicio=? WHERE id=?",
                (nombre_archivo, cliente_id)
            )






        # 🔥 recalcular total
        total = monto + (monto * interes / 100)

        # ✅ actualizar cliente
        cursor.execute("""
            UPDATE clientes
            SET nombre=?, apellidos=?, dni=?, direccion=?, telefono=?,
                monto=?, interes=?, total=?, tipo_pago=?, cuotas=?
            WHERE id=?
        """, (nombre, apellidos, dni, direccion, telefono,
              monto, interes, total, tipo_pago, cuotas, cliente_id))

        # 🔥 BORRAR cronograma viejo

        # 🔥 GENERAR NUEVO cronograma
        from datetime import datetime, timedelta



        conn.commit()
        conn.close()

        return redirect("/")

    # GET
    cursor.execute("SELECT * FROM clientes WHERE id=?", (cliente_id,))
    cliente = cursor.fetchone()
    conn.close()

    return render_template("editar.html", cliente=cliente)





@app.route("/cronograma/<int:cliente_id>")
def ver_cronograma(cliente_id):
    conn = sqlite3.connect("DB.PATH")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM clientes WHERE id = ?", (cliente_id,))
    cliente = cursor.fetchone()
    total_prestamo = cliente["total"]

    cursor.execute("SELECT * FROM cronograma WHERE cliente_id = ?", (cliente_id,))
    cuotas = cursor.fetchall()

    print(cuotas[0].keys())

    hoy = datetime.today().date()
    mora_diaria = 5

    cuotas_actualizadas = []
    total_pagado = 0
    cuotas_pagadas = 0

    for cuota in cuotas:

        fecha_pago = datetime.strptime(cuota["fecha_pago"], "%d/%m/%Y").date()
        estado = cuota["estado"]
        mora_guardada = cuota["mora"] if cuota["mora"] else 0

        mora = mora_guardada  # usamos la que ya está guardada
        puntualidad = "-"

        if not estado.startswith("Pagado"):

            if hoy > fecha_pago:

                dias_atraso = (hoy - fecha_pago).days
                mora_calculada = dias_atraso * MORA_DIARIA

                # SOLO actualizamos si es diferente
                if mora_calculada != mora_guardada:
                    cursor.execute("""
                        UPDATE cronograma
                        SET mora = ?
                        WHERE id = ?
                    """, (mora_calculada, cuota["id"]))
                    conn.commit()

                    mora = mora_calculada

                estado = "Vencido"
                puntualidad = "Retrasado"

            else:
                estado = "Pendiente"
        else:
            if estado == "Pagado con atraso":
                puntualidad = "Retrasado"
            else:
                puntualidad = "Puntual"

        # 🔥 SUMA CORRECTA
        if estado.startswith("Pagado"):
            total_pagado += cuota["cuota"]  # SOLO capital
            cuotas_pagadas += 1

        monto_base = cuota["cuota"]
        monto_total = monto_base + mora

        cuotas_actualizadas.append((
            cuota[0],  # 0 id
            cuota[2],  # 1 fecha
            cuota[3],  # 2 monto base
            mora,  # 3 mora
            cuota[3] + mora,  # 4 monto total
            estado,  # 5 estado
            puntualidad  # 6 puntualidad
        ))

    total_pendiente = total_prestamo - total_pagado
    conn.close()

    return render_template(
        "cronograma.html",
        cuotas=cuotas_actualizadas,
        total_prestamo=total_prestamo,
        total_pagado=total_pagado,
        total_pendiente=total_pendiente,
        cuotas_pagadas=cuotas_pagadas,
        cliente_id=cliente_id,
        cliente=cliente
    )



@app.route('/editar_cronograma/<int:cliente_id>', methods=['GET', 'POST'])
@login_required
def editar_cronograma(cliente_id):

    conn = sqlite3.connect('prestamos.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM clientes WHERE id = ?", (cliente_id,))
    cliente = cursor.fetchone()

    cursor.execute("SELECT * FROM cronograma WHERE cliente_id = ?", (cliente_id,))
    cuotas_db = cursor.fetchall()

    from datetime import datetime

    # LISTA PARA MOSTRAR EN HTML
    cuotas = []

    for c in cuotas_db:
        fecha_obj = datetime.strptime(c['fecha_pago'], "%d/%m/%Y")

        cuotas.append({
            'id': c['id'],
            'fecha_pago': fecha_obj.strftime("%Y-%m-%d"),
            'cuota': c['cuota'],
            'estado': c['estado']
        })

    # ---------------------------------
    # CUANDO SE GUARDAN CAMBIOS
    # ---------------------------------

    if request.method == 'POST':

        for cuota in cuotas_db:

            cuota_id = cuota['id']

            fecha_html = request.form.get(f'fecha_{cuota_id}')
            nuevo_monto = request.form.get(f'monto_{cuota_id}')

            fecha_obj = datetime.strptime(fecha_html, "%Y-%m-%d")
            nueva_fecha = fecha_obj.strftime("%d/%m/%Y")

            fecha_antigua = cuota['fecha_pago']

            # SI CAMBIA LA FECHA -> REINICIAR SOLO ESA CUOTA
            if nueva_fecha != fecha_antigua:
                nuevo_estado = "Pendiente"
            else:
                nuevo_estado = cuota['estado']

            cursor.execute("""
            UPDATE cronograma
            SET fecha_pago = ?, cuota = ?, estado = ?
            WHERE id = ?
            """, (nueva_fecha, nuevo_monto, nuevo_estado, cuota_id))

        conn.commit()
        conn.close()

        return redirect(f'/cronograma/{cliente_id}')

    # ---------------------------------
    # MOSTRAR PANTALLA EDITAR
    # ---------------------------------

    conn.close()
    return render_template('editar_cronograma.html', cliente=cliente, cuotas=cuotas)




@app.route('/actualizar_bd')
def actualizar_bd():
    conn = sqlite3.connect("DB.PATH")
    cursor = conn.cursor()

    try:
        cursor.execute("ALTER TABLE cronograma ADD COLUMN mora REAL DEFAULT 0")
    except:
        pass

    try:
        cursor.execute("ALTER TABLE cronograma ADD COLUMN puntualidad TEXT DEFAULT '-'")
    except:
        pass

    conn.commit()
    conn.close()

    return "Base de datos actualizada"




@app.context_processor
def contar_vencidos():
    hoy = date.today().strftime("%Y-%m-%d")

    conexion = sqlite3.connect('prestamos.db')
    cursor = conexion.cursor()

    cursor.execute("""
        SELECT COUNT(*)
        FROM cronograma
        WHERE fecha_pago < ?
        AND estado = 'Pendiente'
    """, (hoy,))

    cantidad = cursor.fetchone()[0]

    conexion.close()

    return dict(total_vencidos=cantidad)

@app.route('/eliminar_cronograma/<int:cronograma_id>')
@login_required
def eliminar_cronograma(cronograma_id):

    conexion = sqlite3.connect('prestamos.db')
    cursor = conexion.cursor()

    # Eliminamos solo esa cuota
    cursor.execute("DELETE FROM cronograma WHERE id = ?", (cronograma_id,))

    conexion.commit()
    conexion.close()

    return redirect('/reporte')


@app.route('/eliminar/<int:cliente_id>')
@login_required
def eliminar(cliente_id):

    conexion = sqlite3.connect('prestamos.db')
    cursor = conexion.cursor()

    # 1️⃣ Eliminar cronograma del cliente
    cursor.execute("DELETE FROM cronograma WHERE cliente_id = ?", (cliente_id,))

    # Verificar si cronograma quedó vacío
    cursor.execute("SELECT COUNT(*) FROM cronograma")
    cantidad_cronograma = cursor.fetchone()[0]

    if cantidad_cronograma == 0:
        cursor.execute("DELETE FROM sqlite_sequence WHERE name='cronograma'")


    # 2️⃣ Eliminar cliente
    cursor.execute("DELETE FROM clientes WHERE id = ?", (cliente_id,))

    # Verificar si clientes quedó vacío
    cursor.execute("SELECT COUNT(*) FROM clientes")
    cantidad_clientes = cursor.fetchone()[0]

    if cantidad_clientes == 0:
        cursor.execute("DELETE FROM sqlite_sequence WHERE name='clientes'")


    conexion.commit()
    conexion.close()

    return redirect('/reporte')



@app.route('/ver_tablas')
def ver_tablas():
    conexion = sqlite3.connect('prestamos.db')
    cursor = conexion.cursor()

    cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
    tablas = cursor.fetchall()

    conexion.close()

    return str(tablas)




@app.route("/pagar/<int:cuota_id>")
def pagar(cuota_id):

    conn = sqlite3.connect("DB.PATH")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM cronograma WHERE id = ?", (cuota_id,))
    cuota = cursor.fetchone()

    if not cuota:
        conn.close()
        return redirect("/")

    mora = cuota["mora"] if cuota["mora"] else 0

    # Determinar tipo de pago según mora guardada
    if mora > 0:
        estado = "Pagado con atraso"
    else:
        estado = "Pagado puntual"

    cursor.execute("""
        UPDATE cronograma
        SET estado = ?
        WHERE id = ?
    """, (estado, cuota_id))

    conn.commit()
    conn.close()

    # Volver al cronograma del cliente
    return redirect(f"/cronograma/{cuota['cliente_id']}")





from datetime import datetime

@app.route('/reporte', methods=['GET', 'POST'])
@login_required
def reporte():
    cantidad = None
    resultados = []

    if request.method == 'POST':

        fecha_inicio = request.form['fecha_inicio']

        # Convertimos la fecha del input
        fecha_obj = datetime.strptime(fecha_inicio, "%Y-%m-%d")
        mes = fecha_obj.strftime("%m")
        anio = fecha_obj.strftime("%Y")

        conexion = sqlite3.connect('prestamos.db')
        conexion.row_factory = sqlite3.Row
        cursor = conexion.cursor()

        cursor.execute("""
            SELECT 
                cronograma.id AS cronograma_id,
                clientes.id AS cliente_id,
                clientes.nombre,
                clientes.apellidos,
                clientes.telefono,
                clientes.direccion,
                clientes.interes,
                cronograma.fecha_pago,
                cronograma.estado
            FROM cronograma
            JOIN clientes ON cronograma.cliente_id = clientes.id
            WHERE substr(cronograma.fecha_pago, 4, 2) = ?
            AND substr(cronograma.fecha_pago, 7, 4) = ?
        """, (mes, anio))

        resultados = cursor.fetchall()
        cantidad = len(resultados)

        conexion.close()

    return render_template('reporte.html', cantidad=cantidad, resultados=resultados)








@app.route('/ver_columnas')
def ver_columnas():
    conexion = sqlite3.connect('prestamos.db')
    cursor = conexion.cursor()

    cursor.execute("PRAGMA table_info(cronograma);")
    columnas = cursor.fetchall()

    conexion.close()

    return str(columnas)


from datetime import datetime

@app.route("/vencidos")
def vencidos():

    conn = sqlite3.connect("DB.PATH")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("""
        SELECT clientes.nombre, clientes.apellidos, cronograma.fecha_pago, cronograma.estado, clientes.id
        FROM cronograma
        JOIN clientes ON cronograma.cliente_id = clientes.id
    """)

    datos = cursor.fetchall()

    hoy = datetime.today().date()

    vencidos = []

    for fila in datos:

        fecha_pago = datetime.strptime(fila["fecha_pago"], "%d/%m/%Y").date()

        estado = fila["estado"]

        # Si la fecha ya pasó y no está pagado
        if hoy > fecha_pago and not estado.startswith("Pagado"):
            estado = "Vencido"

        # Guardamos el registro con el nuevo estado
        if estado == "Vencido":
            vencidos.append({
                "id": fila["id"],
                "nombre": fila["nombre"],
                "apellidos": fila["apellidos"],
                "fecha_pago": fila["fecha_pago"],
                "estado": estado
            })


        print("HOY:", hoy)
        print("VENCIDOS ENCONTRADOS:", len(vencidos))
        




    conn.close()

    cantidad = len(vencidos)

    return render_template("vencidos.html", vencidos=vencidos, cantidad=cantidad)



from openpyxl import Workbook
from openpyxl.styles import Font
from flask import send_file
import io

from openpyxl import Workbook
from openpyxl.drawing.image import Image

@app.route("/exportar_cronograma/<int:cliente_id>")
def exportar_cronograma(cliente_id):

    conn = sqlite3.connect("DB.PATH")
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM clientes WHERE id=?", (cliente_id,))
    cliente = cursor.fetchone()

    cursor.execute("SELECT * FROM cronograma WHERE cliente_id=?", (cliente_id,))
    cuotas = cursor.fetchall()

    print(cliente.keys())

    wb = Workbook()
    ws = wb.active
    ws.title = "Cronograma"

    fila = 1

    # ESTILOS
    titulo = Font(size=16, bold=True)
    subtitulo = Font(size=13, bold=True)
    negrita = Font(bold=True)

    centrado = Alignment(horizontal="center")
    izquierda = Alignment(horizontal="left")

    borde = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )



    # -----------------------------
    # DATOS DEL CLIENTE
    # -----------------------------

    ws.merge_cells("A1:D1")
    ws["A1"] = "REPORTE DE CRONOGRAMA DE PRÉSTAMO"
    ws["A1"].font = titulo
    ws["A1"].alignment = centrado

    ws["A3"] = "Nombre"
    ws["A4"].font = negrita
    ws["A5"].font = negrita
    ws["A6"].font = negrita
    ws["A7"].font = negrita
    ws["A8"].font = negrita
    ws["A9"].font = negrita

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 20

    ws["B3"] = cliente["nombre"] + " " + cliente["apellidos"]

    ws["A4"] = "DNI"
    ws["B4"] = cliente["dni"]

    ws["A5"] = "Dirección"
    ws["B5"] = cliente["direccion"]

    ws["A6"] = "Teléfono"
    ws["B6"] = cliente["telefono"]

    ws["A7"] = "Monto"
    ws["B7"] = cliente["monto"]

    ws["A8"] = "Interés (%)"
    ws["B8"] = cliente["interes"]

    ws["A9"] = "Total Préstamo"
    ws["B9"] = cliente["total"]

    # -----------------------------
    # CRONOGRAMA
    # -----------------------------

    ws.merge_cells("A11:C11")
    ws["A11"] = "CRONOGRAMA DE PAGOS"
    ws["A11"].font = subtitulo
    ws["A11"].alignment = centrado



    ws["A13"] = "Fecha Pago"
    ws["B13"] = "Monto Cuota"
    ws["C13"] = "Mora"
    ws["D13"] = "Monto Total"
    ws["E13"] = "Estado"


    for col in ["A13", "B13", "C13", "D13", "E13"]:
        ws[col].font = negrita
        ws[col].alignment = centrado
        ws[col].border = borde

    fila = 14




    for cuota in cuotas:


        mora = cuota["mora"] if cuota["mora"] else 0
        total = cuota["cuota"] + mora


        ws.cell(row=fila, column=1).value = cuota["fecha_pago"]
        ws.cell(row=fila, column=2).value = cuota["cuota"]
        ws.cell(row=fila, column=3).value = mora
        ws.cell(row=fila, column=4).value = total
        ws.cell(row=fila, column=5).value = cuota["estado"]

        for col in range(1,6):
            ws.cell(row=fila, column=col).border = borde
            ws.cell(row=fila, column=col).alignment = centrado

        fila += 1

    # -----------------------------
    # IMÁGENES
    # -----------------------------

    fila += 2
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=4)
    ws.cell(row=fila, column=1).value = "DOCUMENTOS DEL CLIENTE"
    ws.cell(row=fila, column=1).font = subtitulo
    ws.cell(row=fila, column=1).alignment = centrado

    carpeta = "static/uploads/"

    fila += 2

    # FOTO ROSTRO
    if cliente["foto_rostro"]:
        img = Image(carpeta + cliente["foto_rostro"])
        img.width = 200
        img.height = 150
        ws.add_image(img, f"A{fila}")
        ws.cell(row=fila+8, column=1).value = "Foto del cliente"

    # DNI FRENTE
    if cliente["dni_frontal"]:
        img2 = Image(carpeta + cliente["dni_frontal"])
        img2.width = 200
        img2.height = 150
        ws.add_image(img2, f"D{fila}")
        ws.cell(row=fila + 8, column=4).value = "DNI frontal"

    fila += 10

    # DNI POSTERIOR
    if cliente["dni_reverso"]:
        img3 = Image(carpeta + cliente["dni_reverso"])
        img3.width = 200
        img3.height = 150
        ws.add_image(img3, f"A{fila}")
        ws.cell(row=fila + 8, column=1).value = "DNI reverso"

    # RECIBO DE SERVICIO
    if cliente["recibo_servicio"]:
        img4 = Image(carpeta + cliente["recibo_servicio"])
        img4.width = 200
        img4.height = 150
        ws.add_image(img4, f"D{fila}")
        ws.cell(row=fila + 8, column=4).value = "Recibo de servicio"

    archivo = f"cronograma_{cliente_id}.xlsx"
    wb.save(archivo)

    conn.close()

    return send_file(archivo, as_attachment=True)



@app.route("/exportar_todos")
def exportar_todos():
    conn = sqlite3.connect("DB.PATH")
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM clientes")
    clientes = cursor.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte General"

    fila = 1

    ws[f"A{fila}"] = "REPORTE GENERAL DE PRÉSTAMOS"
    ws[f"A{fila}"].font = Font(bold=True, size=14)
    fila += 2

    encabezados = [
    "ID",
    "Nombre",
    "DNI",
    "Teléfono",
    "Dirección",
    "Monto",
    "Interés (%)",
    "Total Préstamo",
    "Ganancia (Interés)",
    "Total Pagado",
    "Total Pendiente",
    "Cuotas Pagadas"
]


    ws.append(encabezados)

    for col in range(1, len(encabezados) + 1):
        ws.cell(row=fila, column=col).font = Font(bold=True)

    fila += 1

    for cliente in clientes:
        cliente_id = cliente[0]
        nombre_completo = cliente[1] + " " + cliente[2]
        dni = cliente[3]
        monto = cliente[6]
        total_prestamo = cliente[8]
        telefono = cliente[5]
        direccion = cliente[4]
        interes = cliente[7]


        cursor.execute("SELECT cuota, estado FROM cronograma WHERE cliente_id = ?", (cliente_id,))
        cuotas = cursor.fetchall()

        total_pagado = 0
        cuotas_pagadas = 0

        for cuota in cuotas:
            if cuota[1] == "Pagado":
                total_pagado += cuota[0]
                cuotas_pagadas += 1

        total_pendiente = total_prestamo - total_pagado
        ganancia = total_prestamo - monto

        ws.append([
  	      cliente_id,
 	      nombre_completo,
   	      dni,
	      telefono,
          direccion,
          monto,
          interes,
          total_prestamo,
          ganancia,
          total_pagado,
          total_pendiente,
          cuotas_pagadas
	])


    # Ajustar columnas automáticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(column)].width = max_length + 2

    conn.close()

    nombre_archivo = "reporte_general.xlsx"
    wb.save(nombre_archivo)

    return send_file(nombre_archivo, as_attachment=True)



import os

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port)



